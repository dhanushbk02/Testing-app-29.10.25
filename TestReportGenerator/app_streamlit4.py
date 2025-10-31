# app_streamlit.py
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from PIL import Image
import tempfile
import os
import sys
from datetime import date

# --- Page setup ---
# ✅ Only run set_page_config if this file is the main entry
#if __name__ == "__main__" or "streamlit" not in sys.modules:
    #st.set_page_config(page_title="Pump Test Results", layout="wide", page_icon="💧")

# --- Compatibility patch for old Streamlit versions ---
if not hasattr(st, "data_editor"):
    st.data_editor = st.experimental_data_editor

# --- Company Header ---
col_logo, col_title = st.columns([2, 6])
with col_logo:
    st.image("fpl-logo-3-2048x345.png", width=500)
with col_title:
    st.markdown("<h1 style='margin-bottom:0px;'>Test Reports Generator</h1>", unsafe_allow_html=True)
    st.caption("Motor & Pump Performance Testing")

st.markdown("---")


#st.title("Pump Test Results (Excel & PDF Report Generator)")

# -------------------------
# Utilities
# -------------------------
def read_models_from_file(file_like_or_path):
    try:
        if isinstance(file_like_or_path, str):
            df = pd.read_excel(file_like_or_path, engine="openpyxl")
        else:
            file_like_or_path.seek(0)
            df = pd.read_excel(file_like_or_path, engine="openpyxl")
    except Exception:
        if isinstance(file_like_or_path, str):
            df = pd.read_excel(file_like_or_path, header=None, engine="openpyxl")
        else:
            file_like_or_path.seek(0)
            df = pd.read_excel(file_like_or_path, header=None, engine="openpyxl")

    col_map = {c.strip().upper(): c for c in df.columns}
    if "MODEL" in col_map:
        model_col = col_map["MODEL"]
    else:
        model_col = df.columns[0]
    pipe_col = None
    for cand in ("PIPE SIZE", "PIPESIZE", "SIZE", "PIPE_SIZE"):
        if cand in col_map:
            pipe_col = col_map[cand]
            break

    labels = []
    info = {}
    for idx, row in df.iterrows():
        m = str(row[model_col]).strip()
        p = str(row[pipe_col]).strip() if pipe_col else ""
        label = f"{m}  —  {p}" if p else m
        labels.append(label)
        info[m] = {str(k): ("" if pd.isna(v) else v) for k, v in row.items()}
    return labels, info

def default_test_rows(n=5):
    return pd.DataFrame({
        "SlNo": list(range(1, n+1)),
        "Flow": [0.0]*n,
        "Head": [0.0]*n,
        "Input_kW": [0.0]*n,
        "UV_ohm": [0.0]*n,
        "VW_ohm": [0.0]*n,
        "WU_ohm": [0.0]*n,
        "Ambient_C": [25.0]*n
    })

def convert_flow_to_lpm(flow, unit_flow):
    arr = np.array(flow, dtype=float)
    if unit_flow == "LPM":
        return arr
    else:
        return arr * 16.6666666667  # m3/hr -> LPM

def convert_head_to_m(head, unit_head):
    arr = np.array(head, dtype=float)
    if unit_head == "m":
        return arr
    elif unit_head == "bar":
        return arr * 10.19716213
    elif unit_head == "kg/cm2":
        return arr * 9.80665
    else:
        return arr

def compute_efficiency_pct(flow_lpm, head_m, input_kw):
    flow = np.array(flow_lpm, dtype=float)
    head = np.array(head_m, dtype=float)
    power = np.array(input_kw, dtype=float)
    eff = np.full(flow.shape, np.nan)
    mask = power > 0
    eff[mask] = (0.0001409 * flow[mask] * head[mask]) / power[mask] * 100.0
    return eff

def affinity_convert(df, D_orig, new_D):
    ratio = new_D / D_orig
    out = df.copy()
    out["Flow"] = out["Flow"] * ratio
    out["Head"] = out["Head"] * (ratio**2)
    out["Input_kW"] = out["Input_kW"] * (ratio**3)
    return out

def fig_to_buf(fig):
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    buf.seek(0)
    return buf

def build_excel_bytes(metadata: dict, wind_df: pd.DataFrame, perf_df: pd.DataFrame, per_diameter: bool, diameters: list, D_orig: float, decimals: int):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(list(metadata.items()), columns=["Field", "Value"]).to_excel(writer, sheet_name="Metadata", index=False)
        wind_df.round(decimals).to_excel(writer, sheet_name="Winding_Resistance", index=False)
        perf_df.round(decimals).to_excel(writer, sheet_name="Performance_Results", index=False)
        if per_diameter and diameters:
            for D in diameters:
                df_conv = affinity_convert(perf_df.copy(), D_orig, D)
                df_conv = df_conv.round(decimals)
                sheet_name = f"D{int(D) if float(D).is_integer() else D}"
                writer.book.create_sheet(sheet_name) if sheet_name not in writer.book.sheetnames else None
                df_conv.to_excel(writer, sheet_name=sheet_name, index=False)
    buf.seek(0)
    return buf

def build_pdf_bytes(metadata: dict, wind_df: pd.DataFrame, perf_df: pd.DataFrame, chart_bufs: dict, decimals: int):
    buf = BytesIO()
    tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    doc = SimpleDocTemplate(tmpf.name, pagesize=A4)
    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph("Pump Test Report", styles["Title"]))
    elems.append(Spacer(1, 6))
    meta_html = "<br/>".join([f"<b>{k}:</b> {v}" for k, v in metadata.items()])
    elems.append(Paragraph(meta_html, styles["Normal"]))
    elems.append(Spacer(1, 10))
    for title, buf_img in chart_bufs.items():
        elems.append(Paragraph(title, styles["Heading3"]))
        im = Image.open(buf_img)
        tmpimg = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        im.save(tmpimg.name, format="PNG")
        elems.append(RLImage(tmpimg.name, width=480, height=260))
        elems.append(Spacer(1, 8))
    elems.append(Paragraph("Winding Resistance (rounded)", styles["Heading3"]))
    wdata = [list(wind_df.columns)]
    for row in wind_df.round(decimals).astype(str).values.tolist():
        wdata.append(row)
    wt = Table(wdata, repeatRows=1)
    wt.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("GRID", (0,0), (-1,-1), 0.25, colors.black),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold")
    ]))
    elems.append(wt)
    elems.append(Spacer(1, 8))

    elems.append(Paragraph("Performance Results (rounded)", styles["Heading3"]))
    pdata = [list(perf_df.columns)]
    for row in perf_df.round(decimals).astype(str).values.tolist():
        pdata.append(row)
    pt = Table(pdata, repeatRows=1)
    pt.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("GRID", (0,0), (-1,-1), 0.25, colors.black),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold")
    ]))
    elems.append(pt)

    doc.build(elems)
    with open(tmpf.name, "rb") as f:
        data = f.read()
    os.remove(tmpf.name)
    return BytesIO(data)

# -------------------------
# Model list (auto-load from local Excel)
# -------------------------
models = []
model_info = {}

# Try to load models automatically from local Excel file(s)
local_names = ["List of models.xlsx", "List of models.xlsm", "List of models.xls"]
for name in local_names:
    if os.path.exists(name):
        try:
            models, model_info = read_models_from_file(name)
            st.success(f"Loaded {len(models)} models from {name}")
            break
        except Exception as e:
            st.warning(f"⚠️ Failed to read {name}: {e}")
            continue

# Fallback — if no Excel found
if not models:
    models = ["Model-A  —  25x25", "Model-B  —  50x38"]
    st.warning("No local 'List of models.xlsx' found. Using default model list.")

# -------------------------
# Section 1: Pump Model Selection & Details
# -------------------------
st.subheader("1. Select Pump Model")

# Show only pump models (remove size)
model_names = [m.split("—")[0].strip() for m in models]
selected_model = st.selectbox("Pump Model (searchable)", options=model_names, index=0)

# Get info for selected model from dictionary
selected_info = model_info.get(selected_model, {}) if model_info else {}

# Extract fields safely (case-insensitive + flexible key matching)
def get_field(data, *keys, default=""):
    for k in keys:
        for dk, dv in data.items():
            if str(dk).strip().lower() == str(k).strip().lower():
                return dv
    return default

hp = get_field(selected_info, "HP", "Horse Power")
current = get_field(selected_info, "Current", "Current (A)")
voltage = float(get_field(selected_info, "Voltage", "Volt", default=415))
frequency = float(get_field(selected_info, "Frequency", "Freq", default=50))
drawing_no = get_field(
    selected_info,
    "Drawing No.",
    "Drawing No",
    "Drawing_No",
    "Drawing Number",
    "DWG No",
    "DWG",
    "Drg No",
    "Drg"
)
winding_conn = get_field(selected_info, "Winding Connection", "Connection", "Wdg Connection")

# --- Ensure RPM (Speed) is read correctly from List of models Excel ---
try:
    # Try exact field names first (case-insensitive)
    speed_from_info = get_field(
        selected_info,
        "Speed",
        "Speed (RPM)",
        "RATED SPEED",
        "SPEED IN RPM",
        "SPEED RPM",
        "RPM",
        default=""
    )

    # Fallback: try to find a numeric-looking value that’s within 500–4000 (likely an RPM)
    if (speed_from_info in ("", None)) and isinstance(selected_info, dict):
        for k, v in selected_info.items():
            try:
                val = float(str(v).replace(",", "").strip())
                if 500 <= val <= 4000:
                    speed_from_info = int(round(val))
                    break
            except Exception:
                continue

    # Final assignment
    if speed_from_info not in ("", None, ""):
        try:
            speed = int(round(float(speed_from_info)))
        except Exception:
            speed = str(speed_from_info)
except Exception:
    pass

# Line 1: Model, HP, Current, Speed
st.markdown(
    f"""
    **Model:** {selected_model} &nbsp;&nbsp;&nbsp;
    **HP:** {hp} &nbsp;&nbsp;&nbsp;
    **Current:** {current} A &nbsp;&nbsp;&nbsp;
    **Speed:** {speed} RPM  
    """
)


# --- Size (from List of models Excel via selected_info) ---
pump_size = get_field(
    selected_info,
    "Size",
    "SIZE",
    "PIPE SIZE",
    "PIPESIZE",
    "PIPE_SIZE",
    "IMPELLER OD",
    "IMPELLER_OD",
    default=""
)
try:
    globals()["pump_size"] = pump_size
except Exception:
    pass
st.markdown(f"**Size:** {pump_size} &nbsp;&nbsp;&nbsp;")

# Line 2: Voltage, Frequency, Drawing No., Winding Connection
st.markdown(
    f"""
    **Voltage:** {voltage} V &nbsp;&nbsp;&nbsp;
    **Frequency:** {frequency} Hz &nbsp;&nbsp;&nbsp;
    **Drawing No.:** {drawing_no} &nbsp;&nbsp;&nbsp;
    **Winding Connection:** {winding_conn}
    """
)



# -------------------------
# Section 2: Metadata & settings (with Flow & Head Reference tabs)
# -------------------------
st.subheader("2. Pump Nameplate Data")
c1, c2, c3 = st.columns(3) 
with c1:
    comp_no = st.text_input("Comp. No.")
with c2:
    test_date = st.date_input("Date", value=date.today())
with c3:
    oa_no = st.text_input("OA. No.")

c4, c5, c6 = st.columns(3)

with c4:
    # Flow unit selector removed — default value retained for compatibility
    flow_unit = "LPM"
    # keep layout spacing consistent
    st.markdown("")  

with c5:
    # Head unit selector removed — default value retained for compatibility
    head_unit = "m"
    st.markdown("")

with c6:
    # Original impeller OD input removed — keep variable for compatibility (default 0)
    impeller_od = 0
    st.markdown("")




# Flow, Head & Input Power reference inputs (side by side, compact)
st.markdown("### Flow, Head & Input Power at Duty Point")

# Create 4 equal columns (added Impeller OD in mm on far right)
col1, col2, col3, col4 = st.columns(4)

with col1:
    ref_flow = st.number_input(
        "Duty - Flow (LPM)",
        min_value=0.0,
        value=0.0,
        step=1.0,             # Flow increments by 1 LPM
        format="%.0f",        # No decimal places (integer values)
        key="ref_flow",
        label_visibility="visible"
    )

with col2:
    ref_head = st.number_input(
        "Duty - Head (m)",
        min_value=0.0,
        value=0.0,
        step=0.1,             # Head increments by 0.1 m
        format="%.1f",
        key="ref_head",
        label_visibility="visible"
    )

with col3:
    ref_input_kw = st.number_input(
        "Input Power at Duty Point (kW)",
        min_value=0.00,
        value=0.00,
        step=0.01,            # Power increments by 0.01 kW
        format="%.2f",
        key="ref_input_kw",
        label_visibility="visible"
    )

with col4:
    impeller_od = st.number_input(
        "Impeller OD in mm",
        min_value=0.0,
        value=0.0,
        step=0.1,             # increments of 0.1
        format="%.1f",        # 0.1 decimal place
        key="impeller_od",
        label_visibility="visible"
    )




# -------------------------
# Section 3: Input mode
# -------------------------
st.subheader("3. Test Data Input")
input_mode = st.radio("Input mode", ["Manual input", "Upload & extract"], index=0)

mapped_df = None
if input_mode == "Upload & extract":
    st.info("Upload an Excel/CSV with columns for SlNo, Flow, Head, Input_kW, etc.")
    uploaded_readings = st.file_uploader("Upload readings", type=["xlsx","xls","csv"])
    if uploaded_readings:
        try:
            tmp = pd.read_csv(uploaded_readings) if uploaded_readings.name.endswith(".csv") else pd.read_excel(uploaded_readings)
            st.success("File read — map columns below.")
            cols = list(tmp.columns)
            with st.form("mapping_form"):
                slc = st.selectbox("SlNo", ["(none)"] + cols)
                flowc = st.selectbox("Flow", ["(none)"] + cols)
                headc = st.selectbox("Head", ["(none)"] + cols)
                powerc = st.selectbox("Input kW", ["(none)"] + cols)
                ambc = st.selectbox("Ambient (C)", ["(none)"] + cols)
                submitted = st.form_submit_button("Map & Load")
            if submitted:
                nrows = len(tmp)
                mapped = pd.DataFrame({
                    "SlNo": tmp[slc] if slc != "(none)" else range(1, nrows+1),
                    "Flow": tmp[flowc] if flowc != "(none)" else 0.0,
                    "Head": tmp[headc] if headc != "(none)" else 0.0,
                    "Input_kW": tmp[powerc] if powerc != "(none)" else 0.0,
                    "Ambient_C": tmp[ambc] if ambc != "(none)" else 25.0
                }).fillna(0).reset_index(drop=True)
                mapped_df = mapped
                st.success("Mapped data loaded below.")
        except Exception as e:
            st.error("Failed to read uploaded file.")
            st.exception(e)

# =============================
# 4.0 WINDING RESISTANCE TEST (Refined & Polished Layout)
# =============================
st.markdown("### 4. Winding Resistance Test")

# --- Get type test resistance & reference temperature from Excel (columns J & K) ---
type_test_value = get_field(selected_info, "Type Test Resistance", "Type Test Value", "TypeTestRes", default="")
ref_temp_value = get_field(selected_info, "Reference Temp", "Reference Temperature", "Temp", default=25)

# Convert safely to numeric
try:
    type_test_value = float(type_test_value)
except:
    type_test_value = None

try:
    ref_temp_value = float(ref_temp_value)
except:
    ref_temp_value = 25.0

# --- Display type test reference info neatly on one line ---
col_left, col_right = st.columns([3, 1.5])
with col_left:
    st.markdown(
        f"**Type Test value in Ω** (Reference Temperature: **{ref_temp_value:.1f} °C**)"
    )
with col_right:
    if type_test_value is not None and type_test_value > 0:
        st.markdown(f"<div style='margin-top:4px;'>**Type Test Value:** {type_test_value:.3f} Ω</div>", unsafe_allow_html=True)
    else:
        # Manual entry if not in Excel
        type_test_value = st.number_input(
            "Enter Type Test Resistance (Ω)",
            min_value=0.0,
            step=0.1,
            key="manual_type_test_res"
        )

# --- Measured temperature line ---
st.markdown("_Measured at ambient temperature_")

# --- Ambient + Table layout (compact & aligned) ---
col_temp, col_table = st.columns([0.7, 3.3])

with col_temp:
    ambient_temp = st.number_input(
        "Ambient Temperature (°C)",
        min_value=0.0,
        step=1.0,
        value=float(ref_temp_value),
        key="ambient_temp",
        format="%.0f"
    )

with col_table:
    # --- Resistance data entry table (compact, no blank rows) ---
    columns = ["UV", "VW", "WU"]
    data = [{"UV": 0.000, "VW": 0.000, "WU": 0.000}]
    df_wind = pd.DataFrame(data)

    df_wind = st.data_editor(
        df_wind,
        hide_index=True,
        use_container_width=False,
        width=350,
        height=95,
        key="df_wind",
        num_rows="fixed",
        column_config={
            "UV": st.column_config.NumberColumn("UV", step=0.001, format="%.3f", width="small"),
            "VW": st.column_config.NumberColumn("VW", step=0.001, format="%.3f", width="small"),
            "WU": st.column_config.NumberColumn("WU", step=0.001, format="%.3f", width="small"),
        },
    )

# --- Compute variation and PASS/FAIL logic ---
if df_wind.values.sum() > 0:
    avg_res = df_wind.mean(axis=1).mean()
    st.write(f"**Average Resistance:** {avg_res:.3f} Ω")

    if type_test_value and type_test_value > 0:
        variation_res = ((avg_res - type_test_value) / type_test_value) * 100
        col_avg, col_result = st.columns([3, 1])
        with col_avg:
            st.write(f"**Variation:** {variation_res:+.2f}%")
        with col_result:
            if avg_res <= type_test_value * 1.05:  # within +5%
                st.success("✅ PASS")
            else:
                st.error("❌ FAIL")
else:
    st.info("Enter measured resistance values (UV, VW, WU) to calculate results.")

st.divider()

# --- Save data safely ---
st.session_state["wind_df"] = df_wind




# =============================
# 5.0 INSULATION RESISTANCE TEST (Enhanced)
# =============================
st.markdown("### 5. Insulation Resistance Test")
st.markdown("_Tested at 500 V DC supply_")

col_ir1, col_ir2 = st.columns([0.7, 3.3])
with col_ir1:
    ir_value = st.number_input(
        "Insulation Resistance (MΩ)",
        min_value=0,
        step=1,
        format="%d",
        key="ir_value"
    )

with col_ir2:
    if ir_value > 100:
        st.success("✅ PASS")
    elif ir_value > 0:
        st.error("❌ FAIL")
        st.warning("⚠️ The minimum acceptable value is 100 MΩ")

st.divider()



# =============================
# 6.0 HIGH VOLTAGE BREAKDOWN TEST (Dynamic for Model Change)
# =============================
st.markdown("### 6. High Voltage Breakdown Test")

# --- Safely extract voltage (L) and leakage current (M) from Excel ---
hv_test_value = None
leakage_limit = None

try:
    # Convert to list to access by column order (L = 12th, M = 13th)
    values_list = list(selected_info.values())

    # Column L (index 11)
    if len(values_list) > 11:
        hv_test_value = values_list[11]
    # Column M (index 12)
    if len(values_list) > 12:
        leakage_limit = values_list[12]
except Exception:
    hv_test_value = None
    leakage_limit = None

# Convert values safely
try:
    hv_test_value = float(hv_test_value)
except:
    hv_test_value = 0.0

try:
    leakage_limit = str(leakage_limit)
except:
    leakage_limit = ""

# --- Display reference values ---
st.markdown(
    f"**To apply {hv_test_value:.2f} kV for 1 min**<br>"
    f"**Type Test Voltage (kV):** {hv_test_value:.2f} &nbsp;&nbsp;&nbsp;&nbsp; "
    f"**Allowable Leakage Current (mA):** {leakage_limit}",
    unsafe_allow_html=True
)

# --- Manual entry fallback if no value found ---
if hv_test_value == 0.0:
    hv_test_value = st.number_input(
        "Enter Type Test Voltage (kV)",
        min_value=0.0,
        step=0.1,
        format="%.2f",
        key=f"manual_hv_{selected_model}"
    )

if leakage_limit.strip() in ["", "nan", "None"]:
    leakage_limit = st.text_input(
        "Enter Allowable Leakage Current (mA)",
        key=f"manual_leak_{selected_model}"
    )

# --- Applied voltage entry ---
applied_voltage = st.number_input(
    "Applied Voltage (kV)",
    min_value=0.0,
    step=0.1,
    format="%.2f",
    key=f"applied_voltage_{selected_model}"
)

# --- PASS/FAIL logic ---
if applied_voltage > 0:
    if abs(applied_voltage - hv_test_value) <= 0.05:
        st.success(f"✅ PASS — Applied {applied_voltage:.2f} kV matches required {hv_test_value:.2f} kV")
    else:
        st.error(f"❌ FAIL — Required {hv_test_value:.2f} kV, Applied {applied_voltage:.2f} kV")

st.divider()


# ============================= 
# 7.0 NO-LOAD TEST (fixed session_state key + float dtypes + step)
# =============================
st.markdown("### 7. No-Load Test")

no_load_columns = ["Frequency (Hz)", "RPM", "Voltage (V)", "Current (A)", "Input Power (W)"]
# use floats so data_editor knows these are numeric with decimals
no_load_data = [[0.0, 0.0, 0.0, 0.0, 0.0]]

# create DataFrame with float dtypes explicitly
df_no_load = pd.DataFrame(no_load_data, columns=no_load_columns)
for c in no_load_columns:
    df_no_load[c] = df_no_load[c].astype(float)

# Column config: format + step (important)
try:
    column_config = {
        "Frequency (Hz)": st.column_config.NumberColumn("Frequency (Hz)", step=0.1, format="%.1f"),
        "RPM": st.column_config.NumberColumn("RPM", step=1, format="%.0f"),
        "Voltage (V)": st.column_config.NumberColumn("Voltage (V)", step=1, format="%.0f"),
        "Current (A)": st.column_config.NumberColumn("Current (A)", step=0.1, format="%.1f"),
        "Input Power (W)": st.column_config.NumberColumn("Input Power (W)", step=1, format="%.0f"),
    }
except Exception:
    column_config = {
        "Frequency (Hz)": {"step": 0.1, "format": "%.1f"},
        "RPM": {"step": 1, "format": "%.0f"},
        "Voltage (V)": {"step": 1, "format": "%.0f"},
        "Current (A)": {"step": 0.1, "format": "%.1f"},
        "Input Power (W)": {"step": 1, "format": "%.0f"},
    }

# show editor and capture edits (widget key remains "df_no_load")
edited_no_load = st.data_editor(
    df_no_load,
    num_rows="dynamic",
    use_container_width=True,
    key="df_no_load",
    column_config=column_config
)

# ensure columns remain numeric after edit
for c in no_load_columns:
    edited_no_load[c] = pd.to_numeric(edited_no_load[c], errors="coerce").fillna(0.0)

# OPTIONAL: auto-calc Input Power (W) from Voltage * Current
# If you want automatic calculation, uncomment the next line:
# edited_no_load["Input Power (W)"] = (edited_no_load["Voltage (V)"] * edited_no_load["Current (A)"]).round(0)

# Save the edited frame to a DIFFERENT session_state key to avoid Streamlit API error
st.session_state["df_no_load_saved"] = edited_no_load.copy()

# If you still need a kW display, compute it separately (read-only)
# st.markdown("**Input Power (kW)** (derived, read-only)")
# input_kw_display = (edited_no_load["Input Power (W)"] / 1000.0).round(2)
# st.dataframe(input_kw_display.to_frame("Input Power (kW)"), use_container_width=True)

st.divider()



# ============================= 
# 8.0 LOCKED ROTOR TEST
# =============================
st.markdown("### 8. Locked Rotor Test")

st.info("Test at lower voltage (typically 100 V).")

# Define columns for Locked Rotor Test
locked_rotor_columns = ["Applied Voltage (V)", "Locked Current (A)", "Input Power (W)"]
locked_rotor_data = [[0.0, 0.0, 0.0]]  # single editable row initially

# Editable table for test entry
df_locked_rotor = pd.DataFrame(locked_rotor_data, columns=locked_rotor_columns)
df_locked_rotor = st.data_editor(df_locked_rotor, num_rows="dynamic", use_container_width=True, key="df_locked_rotor")

# --- Extract rated current from model details ---
try:
    rated_current = float(current) if current not in ["", None, "nan"] else 0.0
except:
    rated_current = 0.0

# --- Compute extrapolated current if data entered ---
if not df_locked_rotor.empty and df_locked_rotor["Applied Voltage (V)"].iloc[0] > 0 and df_locked_rotor["Locked Current (A)"].iloc[0] > 0:
    applied_voltage = float(df_locked_rotor["Applied Voltage (V)"].iloc[0])
    locked_current = float(df_locked_rotor["Locked Current (A)"].iloc[0])
    
    # Extrapolate current to 415V
    extrapolated_current = locked_current * (415 / applied_voltage)
    
    # Calculate % of rated current
    if rated_current > 0:
        percent_of_rated = (extrapolated_current / rated_current) * 100
        allowable_current = 6 * rated_current
    else:
        percent_of_rated = 0
        allowable_current = 0
    
    st.markdown(
        f"""
        **Extrapolated current value at 415 V:** {extrapolated_current:.2f} A ({percent_of_rated:.1f}% of rated current)  
        **Allowable current value at 415 V:** {allowable_current:.2f} A (6 × rated current)
        """
    )

    # --- PASS/FAIL logic ---
    if extrapolated_current <= allowable_current:
        st.success("✅ PASS — Locked rotor current is within allowable limit.")
    else:
        st.error("❌ FAIL — Locked rotor current exceeds 6× rated current.")
else:
    st.warning("Enter Data to compute results.")

st.divider()


# -------------------------
# Section 9: Test Results Table (compact) - instant calculation fix
# -------------------------
st.subheader("9. Test Results Table (compact)")
st.info(
    "Enter the electrical and hydraulic test readings below. "
    )

# -------------------------
# Pipe size + orifice constant
# -------------------------
pipe_size = st.selectbox(
    "Select Test Bench Pipe Size:",
    options=["4 inch", "6 inch", "8 inch", "N/A"],
    index=3,
    key="pipe_size_select"
)
standard_orifices = {"4 inch": 70.24, "6 inch": 318.38, "8 inch": 417.44}
if pipe_size == "N/A":
    custom_orifice = st.number_input(
        "Enter custom orifice constant (for N/A):",
        min_value=0.0,
        value=float(st.session_state.get("custom_orifice_input", 0.0)),
        format="%.4f",
        key="custom_orifice_input"
    )
    selected_orifice_constant = float(custom_orifice)
else:
    selected_orifice_constant = float(standard_orifices.get(pipe_size, 0.0))
st.write(f"**Orifice Constant:** {selected_orifice_constant:.4f}")

# -------------------------
# Default table creation (once)
# -------------------------
def default_compact_perf():
    return pd.DataFrame({
        "SlNo": [1, 2, 3, 4, 5],
        "Voltage (V)": [0]*5,
        "Current (A)": [0.0]*5,
        "Input (W)": [0]*5,
        "Head (m)": [0.0]*5,
        "Differential Pressure (mmHg)": [0.0]*5,
        "Flow (LPM)": [0.0]*5,
        "Efficiency (%)": [0.0]*5
    })

if "perf_df" not in st.session_state:
    st.session_state.perf_df = default_compact_perf()

# -------------------------
# Helper function: Flow + Efficiency calculation
# -------------------------
def compute_flow_efficiency(df, orifice_const):
    df = df.copy()
    df["Differential Pressure (mmHg)"] = pd.to_numeric(df["Differential Pressure (mmHg)"], errors="coerce").fillna(0.0)
    df["Head (m)"] = pd.to_numeric(df["Head (m)"], errors="coerce").fillna(0.0)
    df["Input (W)"] = pd.to_numeric(df["Input (W)"], errors="coerce").fillna(0.0)

    # Flow calculation
    if orifice_const > 0:
        df["Flow (LPM)"] = df["Differential Pressure (mmHg)"].apply(
            lambda dp: (dp ** 0.5) * orifice_const if dp > 0 else 0.0
        )
    else:
        df["Flow (LPM)"] = 0.0

    # Efficiency calculation (%)
    def calc_eff(row):
        input_kw = row["Input (W)"] / 1000.0
        if input_kw == 0:
            return 0.0
        eff = (0.0001409 * row["Flow (LPM)"] * row["Head (m)"]) / input_kw
        return eff * 100.0  # convert to percent

    df["Efficiency (%)"] = df.apply(calc_eff, axis=1)
    return df

# -------------------------
# Editable table (live)
# -------------------------
try:
    column_config = {
        "Flow (LPM)": st.column_config.NumberColumn("Flow (LPM)", disabled=True),
        "Efficiency (%)": st.column_config.NumberColumn("Efficiency (%)", disabled=True)
    }
except Exception:
    column_config = {
        "Flow (LPM)": {"disabled": True},
        "Efficiency (%)": {"disabled": True}
    }

edited_df = st.data_editor(
    st.session_state.perf_df,
    num_rows="fixed",
    use_container_width=True,
    key="perf_table_small",
    column_config=column_config,
    on_change=lambda: st.session_state.update({"latest_perf_edit": st.session_state.perf_df.copy()})
)

# Keep latest edits available
st.session_state.latest_perf_edit = edited_df.copy()

# -------------------------
# Buttons
# -------------------------
c1, c2 = st.columns([1, 1])
with c1:
    save_calc = st.button("Calculate Flow & Efficiency")
with c2:
    reset_table = st.button("Reset table to default")

# -------------------------
# Actions
# -------------------------
if save_calc:
    edited_df = st.session_state.get("latest_perf_edit", st.session_state.perf_df)
    new_df = compute_flow_efficiency(edited_df, float(selected_orifice_constant))
    st.session_state.perf_df = new_df.copy()
    st.success("✅ Flow and Efficiency updated successfully!")

if reset_table:
    st.session_state.perf_df = default_compact_perf()
    st.session_state.perf_history = []
    st.success("🔄 Table reset to default.")


# -------------------------
# Comparison vs manual reference (duty-point based) — ref input treated as kW, converted to W
# -------------------------
st.markdown("---")
st.markdown("**Comparison vs Manual Reference (duty point based)**")

# tolerances
flow_tol_pct = 6.0   # ±6% for flow
kw_tol_pct = 8.0     # ±8% for input power

# Reference values (expected in session_state)
ref_flow_val = st.session_state.get("ref_flow", ref_flow if 'ref_flow' in locals() else 0.0)
ref_head_val = st.session_state.get("ref_head", ref_head if 'ref_head' in locals() else 0.0)

# Reference input power: explicit kW key (we treat it as kW and convert to W)
# Accept either 'ref_kw' or 'ref_input_kw' as the key names (kW)
ref_kw_val = st.session_state.get("ref_kw",
                 st.session_state.get("ref_input_kw",
                 (ref_kw if 'ref_kw' in locals() else (ref_input_kw if 'ref_input_kw' in locals() else 0.0))))

# convert reference kW -> W
try:
    ref_input_w = float(ref_kw_val) * 1000.0 if ref_kw_val is not None else 0.0
except Exception:
    ref_input_w = 0.0

# Use the persisted (saved) table for comparisons if present; else use preview/edited
perf_for_calc = st.session_state.get("perf_df", preview_df.copy() if 'preview_df' in locals() else edited_df.copy())

# Ensure numeric columns exist
perf_for_calc["Head (m)"] = pd.to_numeric(perf_for_calc.get("Head (m)", 0.0), errors="coerce").fillna(0.0)
perf_for_calc["Flow (LPM)"] = pd.to_numeric(perf_for_calc.get("Flow (LPM)", 0.0), errors="coerce").fillna(0.0)
if "Input (W)" in perf_for_calc.columns:
    perf_for_calc["Input (W)"] = pd.to_numeric(perf_for_calc["Input (W)"], errors="coerce").fillna(0.0)
else:
    perf_for_calc["Input (W)"] = 0.0

# Find duty-row: row whose Head is closest to reference head
if ref_head_val and float(ref_head_val) != 0.0:
    abs_diff = (perf_for_calc["Head (m)"] - float(ref_head_val)).abs()
    duty_idx = int(abs_diff.idxmin())
else:
    duty_idx = 0

# Extract measured values at duty row
meas_flow = float(perf_for_calc.at[duty_idx, "Flow (LPM)"])
meas_input_w = float(perf_for_calc.at[duty_idx, "Input (W)"])
meas_head = float(perf_for_calc.at[duty_idx, "Head (m)"])

st.write(f"**Duty point used for comparison:** row #{duty_idx+1} — Head = {meas_head:.3f} m, Flow = {meas_flow:.3f} LPM, Input = {meas_input_w:.2f} W")

# Flow comparison
if ref_flow_val and float(ref_flow_val) != 0.0:
    flow_variation = (meas_flow - float(ref_flow_val)) / float(ref_flow_val) * 100.0
    st.write(f"Flow Reference: **{float(ref_flow_val):.3f} LPM** — Measured: **{meas_flow:.3f} LPM** — Variation: **{flow_variation:+.2f}%** (Tol ±{flow_tol_pct}%)")
    if abs(flow_variation) <= flow_tol_pct:
        st.success("✅ Flow PASS")
    else:
        st.error("❌ Flow FAIL")
else:
    st.info("Flow Reference: **—** (enter in Section 2)")

# Input (kW -> W) comparison
if ref_input_w and float(ref_input_w) != 0.0:
    kw_variation = (meas_input_w - ref_input_w) / ref_input_w * 100.0
    st.write(
        f"Input Power Reference: **{ref_input_w:.2f} W** (from {ref_kw_val} kW) — "
        f"Measured (table): **{meas_input_w:.2f} W** — "
        f"Variation: **{kw_variation:+.2f}%** (Ref. ≤ value considered PASS)"
    )

    if meas_input_w > ref_input_w:
        st.error("❌ Input Power FAIL — Measured value exceeds reference.")
    else:
        st.success(f"✅ Input Power PASS — Measured value is lower by {abs(kw_variation):.2f}%")
else:
    st.info("Input Power Reference: **—** (enter reference kW in Section 2)")



# -------------------------
# 10. Pump Performance Curves (three stacked plots, shared X scale + 15% allowance)
# -------------------------
st.subheader("10. Pump Performance Curves")
show_charts = st.checkbox("Show performance plots", value=True)
chart_bufs = {}

# safe perf_df alias
perf_df = st.session_state.get("perf_df", None)
if perf_df is None:
    if 'preview_df' in globals():
        perf_df = preview_df.copy()
    elif 'edited_df' in globals():
        perf_df = edited_df.copy()
    else:
        try:
            perf_df = default_compact_perf()
        except Exception:
            perf_df = pd.DataFrame({
                "SlNo": [1,2,3,4,5],
                "Flow (LPM)": [0.0]*5,
                "Head (m)": [0.0]*5,
                "Input (W)": [0.0]*5,
                "Differential Pressure (mmHg)": [0.0]*5
            })

# ensure numeric columns
perf_df["Flow (LPM)"] = pd.to_numeric(perf_df.get("Flow (LPM)", 0.0), errors="coerce").fillna(0.0)
perf_df["Head (m)"] = pd.to_numeric(perf_df.get("Head (m)", 0.0), errors="coerce").fillna(0.0)
perf_df["Input (W)"] = pd.to_numeric(perf_df.get("Input (W)", 0.0), errors="coerce").fillna(0.0)

# compute Efficiency (%) if missing or NaN
if "Efficiency (%)" not in perf_df.columns or perf_df["Efficiency (%)"].isna().any():
    def _calc_eff(r):
        in_w = r["Input (W)"]
        in_kw = in_w / 1000.0 if in_w else 0.0
        if in_kw == 0:
            return 0.0
        eff = (0.0001409 * r["Flow (LPM)"] * r["Head (m)"]) / in_kw
        return eff * 100.0
    perf_df["Efficiency (%)"] = perf_df.apply(_calc_eff, axis=1)

# sort by Flow for smooth curves
plot_df = perf_df.copy().sort_values("Flow (LPM)").reset_index(drop=True)
x = plot_df["Flow (LPM)"].astype(float).values
y_head = plot_df["Head (m)"].astype(float).values
y_eff = plot_df["Efficiency (%)"].astype(float).values
y_input_kw = (plot_df["Input (W)"].astype(float).values) / 1000.0  # kW

if show_charts:
    try:
        # --- compute shared X limit (15% allowance) and round to nearest 100 ---
        x_max = max(x.max() if len(x)>0 else 0.0, 0.0)
        x_allow = x_max * 1.15
        # round up to nearest 100 (choose sensible base)
        if x_allow <= 100:
            x_round_base = 10
        else:
            x_round_base = 100
        x_limit = int(np.ceil(x_allow / x_round_base) * x_round_base)

        # --- compute Y limits with 15% allowance and round to sensible ticks ---
        # Head
        h_max = max(y_head.max() if len(y_head)>0 else 0.0, 0.0)
        h_allow = h_max * 1.15
        # round head to nearest 0.5 if small, else 1
        h_round_base = 0.5 if h_allow <= 10 else 1.0
        h_limit = float(np.ceil(h_allow / h_round_base) * h_round_base)

        # Efficiency (%) - cap at 100 but allow 15% extra then clamp
        e_max = max(y_eff.max() if len(y_eff)>0 else 0.0, 0.0)
        e_allow = e_max * 1.15
        # round efficiency to nearest 1
        e_limit = float(np.ceil(e_allow / 1.0) * 1.0)
        if e_limit > 100.0:
            e_limit = 100.0

        # Input (kW)
        p_max = max(y_input_kw.max() if len(y_input_kw)>0 else 0.0, 0.0)
        p_allow = p_max * 1.15
        # choose rounding: if small, 0.1 else 0.5 or 1.0
        if p_allow <= 1:
            p_round = 0.1
        elif p_allow <= 5:
            p_round = 0.5
        else:
            p_round = 1.0
        p_limit = float(np.ceil(p_allow / p_round) * p_round)

        # --- Plot 1: Flow vs Head ---
        fig_h, ax_h = plt.subplots(figsize=(6,3))
        ax_h.plot(x, y_head, marker="o", linestyle="-", linewidth=1.5, color="black")
        ax_h.set_xlabel("Flow (LPM)")
        ax_h.set_ylabel("Head (m)")
        ax_h.set_title("Flow vs Head")
        ax_h.grid(True, linestyle="--", linewidth=0.5)
        ax_h.set_xlim(0, x_limit)
        ax_h.set_ylim(0, h_limit if h_limit>0 else 1)
        fig_h.tight_layout()
        st.pyplot(fig_h)
        chart_bufs["Flow_vs_Head"] = fig_to_buf(fig_h)
        plt.close(fig_h)

        # --- Plot 2: Flow vs Efficiency (%) ---
        fig_e, ax_e = plt.subplots(figsize=(6,3))
        ax_e.plot(x, y_eff, marker="^", linestyle="-", linewidth=1.5, color="tab:green")
        ax_e.set_xlabel("Flow (LPM)")
        ax_e.set_ylabel("Efficiency (%)")
        ax_e.set_title("Flow vs Efficiency")
        ax_e.grid(True, linestyle="--", linewidth=0.5)
        ax_e.set_xlim(0, x_limit)
        ax_e.set_ylim(0, e_limit if e_limit>0 else 10)
        fig_e.tight_layout()
        st.pyplot(fig_e)
        chart_bufs["Flow_vs_Efficiency"] = fig_to_buf(fig_e)
        plt.close(fig_e)

        # --- Plot 3: Flow vs Input (kW) ---
        fig_i, ax_i = plt.subplots(figsize=(6,3))
        ax_i.plot(x, y_input_kw, marker="s", linestyle="--", linewidth=1.5, color="tab:orange")
        ax_i.set_xlabel("Flow (LPM)")
        ax_i.set_ylabel("Input (kW)")
        ax_i.set_title("Flow vs Input")
        ax_i.grid(True, linestyle="--", linewidth=0.5)
        ax_i.set_xlim(0, x_limit)
        ax_i.set_ylim(0, p_limit if p_limit>0 else 1)
        fig_i.tight_layout()
        st.pyplot(fig_i)
        chart_bufs["Flow_vs_Input_kW"] = fig_to_buf(fig_i)
        plt.close(fig_i)

    except Exception as e:
        st.warning(f"Unable to draw charts: {e}")


# -------------------------
# Section: Export using Template (single sheet, no charts) - FIXED
# -------------------------
from openpyxl import load_workbook
from io import BytesIO

TEMPLATE_PATH = "Certificate_Template.xlsx"  # keep in same folder as app_streamlit.py

def _find_cell_with_text(ws, text):
    """Return (row, col) of the first cell that contains given text (case-insensitive)."""
    text_low = str(text).strip().lower()
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and text_low in cell.value.strip().lower():
                return cell.row, cell.column
    return None

def _find_header_row(ws, start_row, header_keywords):
    """Find header row by scanning below the given start_row."""
    for r in range(start_row, start_row + 6):
        header_map = {}
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        joined = " | ".join([str(x).lower() if x else "" for x in row_vals])
        if any(k.lower() in joined for k in header_keywords):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v and isinstance(v, str) and v.strip():
                    header_map[v.strip().lower()] = c
            return r, header_map
    return None, {}

def _safe_row_get(row, keys, default=0):
    """Try keys in order and return first non-empty value; else default."""
    for k in keys:
        if k in row and row.get(k) not in (None, ""):
            return row.get(k)
    return default

def generate_certificate_from_template_bytes(template_path: str,
                                              metadata: dict,
                                              perf_df: 'pd.DataFrame',
                                              decimals: int = 2) -> BytesIO:
    """Load the Excel template, fill metadata & performance table, return BytesIO."""
    wb = load_workbook(template_path)
    ws = wb.active  # assume main certificate sheet

    # === 1) Write header details to specific cells per your mapping ===
    try:
        if metadata.get("date", "") != "":
            ws["F2"].value = metadata.get("date")            # F2: DATE
        if metadata.get("oa_no", "") != "":
            ws["F3"].value = metadata.get("oa_no")           # F3: OA. No.

        ws["B2"].value = metadata.get("type", "")           # B2: Pump Type / Model
        ws["B3"].value = metadata.get("size", "")           # B3: Pump Size
        ws["B4"].value = metadata.get("volts", "")          # B4: Voltage (if any)
        ws["B5"].value = metadata.get("rpm", "")            # B5: Speed (RPM)

        # === Duty display (B6 merged cell B6:C6:D6) ===
        try:
            # Try to get flow & head from several likely places (globals, session_state, perf_df, metadata)
            flow_val = None
            head_val = None
            try:
                import streamlit as _st
                if "ref_flow" in _st.session_state:
                    flow_val = _st.session_state.get("ref_flow")
                if "ref_head" in _st.session_state:
                    head_val = _st.session_state.get("ref_head")
            except Exception:
                pass

            # fallback to globals (what your app often uses)
            if flow_val in (None, "") and "ref_flow" in globals():
                flow_val = globals().get("ref_flow")
            if head_val in (None, "") and "ref_head" in globals():
                head_val = globals().get("ref_head")

            # fallback to perf_df first row (if still missing)
            try:
                if (flow_val in (None, "") or head_val in (None, "")) and hasattr(perf_df, "iloc") and len(perf_df) > 0:
                    first_pf = perf_df.iloc[0]
                    if flow_val in (None, ""):
                        flow_val = first_pf.get("Flow (LPM)", first_pf.get("Flow", first_pf.get("LPM", flow_val)))
                    if head_val in (None, ""):
                        head_val = first_pf.get("Head (m)", first_pf.get("Head", head_val))
            except Exception:
                pass

            # fallback to metadata 'duty'
            if (flow_val in (None, "")) or (head_val in (None, "")):
                duty_meta = metadata.get("duty", "")
                if duty_meta and isinstance(duty_meta, str) and "@" in duty_meta:
                    ws["B6"].value = duty_meta
                else:
                    if flow_val in (None, ""):
                        flow_val = metadata.get("ref_flow", metadata.get("flow", flow_val))
                    if head_val in (None, ""):
                        head_val = metadata.get("ref_head", metadata.get("head", head_val))

            # Write duty string if values found
            if flow_val not in (None, "") and head_val not in (None, ""):
                try:
                    fnum = int(round(float(flow_val)))
                    hnum = int(round(float(head_val)))
                    ws["B6"].value = f"{fnum} Lpm @ {hnum} m"
                    try:
                        ws.merge_cells("B6:D6")
                    except Exception:
                        pass
                except Exception:
                    try:
                        ws["B6"].value = f"{flow_val} Lpm @ {head_val} m"
                        try: ws.merge_cells("B6:D6")
                        except: pass
                    except Exception:
                        pass
        except Exception:
            pass

        # Keep A7, C7 & E7 as-is (do not overwrite)


        if metadata.get("connection", "") != "":
            ws["F6"].value = metadata.get("connection")     # F6: Winding connection
        if metadata.get("drawing_no", "") != "":
            ws["F7"].value = metadata.get("drawing_no")    # F7: Drawing number

        if metadata.get("comp_no", "") != "":
            ws["D2"].value = metadata.get("comp_no")        # D2: Component / Serial No
        if metadata.get("amps", "") != "":
            val_amps = metadata.get("amps")
            try:
                ws["D3"].value = round(float(val_amps), 1)   # D3: Amps (1 decimal)
            except Exception:
                ws["D3"].value = val_amps
        if metadata.get("hp", "") != "":
            ws["D4"].value = metadata.get("hp")             # D4: HP
        if metadata.get("hz", "") != "":
            try:
                ws["D5"].value = round(float(metadata.get("hz")), 1)
            except Exception:
                ws["D5"].value = metadata.get("hz")

    except Exception:
        pass

    # === 1b) Ambient temp & winding resistances (D10, F10, G10, H10) ===
    try:
        if metadata.get("ambient_temp", "") != "":
            ws["D10"].value = metadata.get("ambient_temp")   # D10: Ambient Temperature (°C)
        # Winding resistances raw values (no rounding)
        if metadata.get("res_uv", "") != "":
            ws["F10"].value = metadata.get("res_uv")        # F10: UV
        if metadata.get("res_vw", "") != "":
            ws["G10"].value = metadata.get("res_vw")        # G10: VW
        if metadata.get("res_wu", "") != "":
            ws["H10"].value = metadata.get("res_wu")        # H10: WU
    except Exception:
        pass

    # === 2) High voltage (C12) ===
    try:
        hv = metadata.get("hv_test", "")
        if hv != "" and hv is not None:
            try:
                hv_num = float(hv) * 1000.0  # multiply kV -> V
                ws["C12"].value = int(round(hv_num))
            except Exception:
                try:
                    ws["C12"].value = int(round(float(hv)))
                except Exception:
                    ws["C12"].value = hv
    except Exception:
        pass

    # === 3) No-load test row (C14..G14): Freq, Speed, Voltage, Current, Power ===
    try:
        nl = metadata.get("no_load", {})  # expect a dict with keys freq,rpm,volt,curr,power
        if isinstance(nl, dict):
            if nl.get("freq", "") != "":
                try: ws["C14"].value = round(float(nl.get("freq")), 1)
                except: ws["C14"].value = nl.get("freq")
            if nl.get("rpm", "") != "":
                try: ws["D14"].value = int(round(float(nl.get("rpm"))))
                except: ws["D14"].value = nl.get("rpm")
            if nl.get("volt", "") != "":
                try: ws["E14"].value = int(round(float(nl.get("volt"))))
                except: ws["E14"].value = nl.get("volt")
            if nl.get("curr", "") != "":
                try: ws["F14"].value = round(float(nl.get("curr")), 1)
                except: ws["F14"].value = nl.get("curr")
            if nl.get("power", "") != "":
                try: ws["G14"].value = int(round(float(nl.get("power"))))
                except: ws["G14"].value = nl.get("power")
    except Exception:
        pass

    # === 4) Locked rotor test (E16,F16,G16) Voltage, Current, Watts ===
    try:
        # helper converters
        def _to_int_safe(v):
            try:
                return int(round(float(v)))
            except Exception:
                return None
        def _to_float1_safe(v):
            try:
                return round(float(v), 1)
            except Exception:
                return None

        # 1) prefer session_state df_locked_rotor
        lr_candidate = None
        try:
            import streamlit as _st
            if "df_locked_rotor" in _st.session_state:
                lr_candidate = _st.session_state.get("df_locked_rotor")
        except Exception:
            lr_candidate = None

        # 2) fallback to globals
        if lr_candidate is None and "df_locked_rotor" in globals():
            lr_candidate = globals().get("df_locked_rotor")

        # 3) fallback to metadata dict
        if lr_candidate is None:
            lr_meta = metadata.get("locked_rotor", None)
            if isinstance(lr_meta, dict) and any(lr_meta.get(k) not in (None, "") for k in ("volt", "curr", "power")):
                lr_candidate = lr_meta

        # normalize to a first-row mapping
        first_row = None
        if lr_candidate is not None:
            try:
                # pandas DataFrame / Series
                if hasattr(lr_candidate, "iloc"):
                    if len(lr_candidate) > 0:
                        first_row = lr_candidate.iloc[0]
                # list-of-dicts
                elif isinstance(lr_candidate, list) and len(lr_candidate) > 0 and isinstance(lr_candidate[0], dict):
                    first_row = lr_candidate[0]
                # dict-like: could be {col: [values]} or {col: scalar}
                elif isinstance(lr_candidate, dict):
                    # detect dict-of-lists (columns -> lists)
                    is_col_lists = False
                    for vv in lr_candidate.values():
                        if isinstance(vv, (list, tuple, pd.Series)):
                            is_col_lists = True
                            break
                    if is_col_lists:
                        # build a first-row dict by taking first element of each column list (if present)
                        fr = {}
                        for k, vv in lr_candidate.items():
                            try:
                                if isinstance(vv, (list, tuple, pd.Series)) and len(vv) > 0:
                                    fr[k] = vv[0]
                                else:
                                    fr[k] = vv
                            except Exception:
                                fr[k] = vv
                        first_row = fr
                    else:
                        # simple dict mapping column -> value (one-row)
                        first_row = lr_candidate
            except Exception:
                first_row = None

        # helper to get from first_row with multiple possible keys
        def _get_from_first(keys):
            if first_row is None:
                return None
            for k in keys:
                try:
                    # dict-like
                    if isinstance(first_row, dict) and k in first_row:
                        v = first_row.get(k)
                        if v not in (None, ""):
                            return v
                    else:
                        # pandas Series or object with .get
                        try:
                            v = first_row.get(k)
                            if v not in (None, ""):
                                return v
                        except Exception:
                            # attribute access or key access fallback
                            try:
                                v = getattr(first_row, k)
                                if v not in (None, ""):
                                    return v
                            except Exception:
                                pass
                except Exception:
                    pass
            return None

        # try common names
        volt_val = _get_from_first(["Applied Voltage (V)", "Applied Voltage", "Voltage (V)", "Voltage", "Volt"])
        curr_val = _get_from_first(["Locked Current (A)", "Locked Current", "Current (A)", "Current", "Amps"])
        power_val = _get_from_first(["Input Power (W)", "Input Power", "Power (W)", "Power"])

        # write only when present (do not overwrite if missing)
        if volt_val not in (None, ""):
            v_int = _to_int_safe(volt_val)
            ws["E16"].value = v_int if v_int is not None else volt_val
        if curr_val not in (None, ""):
            v_f1 = _to_float1_safe(curr_val)
            ws["F16"].value = v_f1 if v_f1 is not None else curr_val
        if power_val not in (None, ""):
            v_int2 = _to_int_safe(power_val)
            ws["G16"].value = v_int2 if v_int2 is not None else power_val

    except Exception:
        # non-fatal — leave template cells unchanged on error
        pass



    # === 5) Write performance table if a "Performance Test Results" section exists in template ===
    try:
        header_pos = _find_cell_with_text(ws, "performance test results")
        if header_pos:
            header_row_candidate, header_map = _find_header_row(
                ws, header_pos[0] + 1,
                header_keywords=["voltage", "input", "head", "lpm", "amps", "flow"]
            )
            if header_row_candidate:
                start_row = header_row_candidate + 1

                def _col_lookup(pref_names):
                    for name in pref_names:
                        if name.lower() in header_map:
                            return header_map[name.lower()]
                    return None

                volt_col = _col_lookup(["voltage (v)", "voltage", "volt", "v"])
                amps_col = _col_lookup(["amps", "current (a)", "current"])
                input_col = _col_lookup(["input (w)", "input-w", "input"])
                head_col = _col_lookup(["head (m)", "head"])
                lpm_col = _col_lookup(["flow (lpm)", "lpm", "flow"])

                if not header_map:
                    volt_col, amps_col, input_col, head_col, lpm_col = 1, 2, 3, 4, 5

                for i in range(len(perf_df)):
                    r = start_row + i
                    row = perf_df.iloc[i]

                    # Voltage -> Integer
                    val_volt = _safe_row_get(row, ["Voltage (V)", "Voltage", "Volt"], default=None)
                    if volt_col and val_volt not in (None, ""):
                        try:
                            ws.cell(r, volt_col).value = int(round(float(val_volt)))
                        except:
                            ws.cell(r, volt_col).value = val_volt

                    # Current -> 1 decimal
                    val_curr = _safe_row_get(row, ["Current (A)", "Current", "Amps"], default=None)
                    if amps_col and val_curr not in (None, ""):
                        try:
                            ws.cell(r, amps_col).value = round(float(val_curr), 1)
                        except:
                            ws.cell(r, amps_col).value = val_curr

                    # Input -> Integer
                    val_input = _safe_row_get(row, ["Input (W)", "Input", "Power (W)", "Power"], default=None)
                    if input_col and val_input not in (None, ""):
                        try:
                            ws.cell(r, input_col).value = int(round(float(val_input)))
                        except:
                            ws.cell(r, input_col).value = val_input

                    # Head -> Integer
                    val_head = _safe_row_get(row, ["Head (m)", "Head"], default=None)
                    if head_col and val_head not in (None, ""):
                        try:
                            ws.cell(r, head_col).value = int(round(float(val_head)))
                        except:
                            ws.cell(r, head_col).value = val_head

                    # Flow/LPM -> Integer
                    val_flow = _safe_row_get(row, ["Flow (LPM)", "Flow", "LPM"], default=None)
                    if lpm_col and val_flow not in (None, ""):
                        try:
                            ws.cell(r, lpm_col).value = int(round(float(val_flow)))
                        except:
                            ws.cell(r, lpm_col).value = val_flow

                # Write fixed cells C21..C26 and D21..D26 from perf_df if present
                try:
                    if "Input (W)" in perf_df.columns and "Head (m)" in perf_df.columns:
                        for idx in range(6):
                            target_row = 21 + idx
                            try:
                                val_in = perf_df.iloc[idx].get("Input (W)")
                                if val_in not in (None, ""):
                                    ws[f"C{target_row}"].value = int(round(float(val_in)))
                            except Exception:
                                try:
                                    ws[f"C{target_row}"].value = perf_df.iloc[idx].get("Input (W)")
                                except Exception:
                                    pass
                            try:
                                val_h = perf_df.iloc[idx].get("Head (m)")
                                if val_h not in (None, ""):
                                    ws[f"D{target_row}"].value = int(round(float(val_h)))
                            except Exception:
                                try:
                                    ws[f"D{target_row}"].value = perf_df.iloc[idx].get("Head (m)")
                                except Exception:
                                    pass
                except Exception:
                    pass
    except Exception:
        pass

    # === 6) Save and return BytesIO ===
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------- Streamlit UI button ----------
import streamlit as st
import pandas as pd

st.markdown("---")
st.subheader("Generate Certificate (from Template)")

# Build metadata_safe (same as before)
metadata_safe = {
    "type": selected_model if 'selected_model' in globals() else "",
    "size": pump_size if 'pump_size' in globals() else "",
    "volts": voltage if 'voltage' in globals() else "",
    "rpm": speed if 'speed' in globals() else "",
    "duty": f"{ref_flow} LPM @ {ref_head} m" if 'ref_flow' in globals() and 'ref_head' in globals() else "",
    "comp_no": comp_no if 'comp_no' in globals() else "",
    "oa_no": oa_no if 'oa_no' in globals() else "",
    "date": test_date.strftime("%Y-%m-%d") if 'test_date' in globals() else "",
    "amps": current if 'current' in globals() else "",
    "hp": hp if 'hp' in globals() else "",
    "hz": frequency if 'frequency' in globals() else "",
    "connection": winding_conn if 'winding_conn' in globals() else "",
    "drawing_no": drawing_no if 'drawing_no' in globals() else "",
    "ambient_temp": ambient_temp if 'ambient_temp' in globals() else "",
    "res_uv": df_wind["UV"].iloc[0] if 'df_wind' in globals() and "UV" in df_wind.columns and not df_wind.empty else "",
    "res_vw": df_wind["VW"].iloc[0] if 'df_wind' in globals() and "VW" in df_wind.columns and not df_wind.empty else "",
    "res_wu": df_wind["WU"].iloc[0] if 'df_wind' in globals() and "WU" in df_wind.columns and not df_wind.empty else "",
    "hv_test": hv_test_value if 'hv_test_value' in globals() else "",
    "no_load": {
        "freq": (st.session_state.get("df_no_load_saved", pd.DataFrame()).get("Frequency (Hz)", [0])[0])
            if "df_no_load_saved" in st.session_state else "",
        "rpm": (st.session_state.get("df_no_load_saved", pd.DataFrame()).get("RPM", [0])[0])
            if "df_no_load_saved" in st.session_state else "",
        "volt": (st.session_state.get("df_no_load_saved", pd.DataFrame()).get("Voltage (V)", [0])[0])
            if "df_no_load_saved" in st.session_state else "",
        "curr": (st.session_state.get("df_no_load_saved", pd.DataFrame()).get("Current (A)", [0])[0])
            if "df_no_load_saved" in st.session_state else "",
        "power": (st.session_state.get("df_no_load_saved", pd.DataFrame()).get("Input Power (W)", [0])[0])
            if "df_no_load_saved" in st.session_state else "",
    },
    "locked_rotor": {
        "volt": (st.session_state.get("df_locked_rotor", pd.DataFrame()).get("Applied Voltage (V)", [0])[0])
            if "df_locked_rotor" in st.session_state else "",
        "curr": (st.session_state.get("df_locked_rotor", pd.DataFrame()).get("Locked Current (A)", [0])[0])
            if "df_locked_rotor" in st.session_state else "",
        "power": (st.session_state.get("df_locked_rotor", pd.DataFrame()).get("Input Power (W)", [0])[0])
            if "df_locked_rotor" in st.session_state else "",
    },
}

perf_for_export = st.session_state.get("perf_df", perf_df if 'perf_df' in globals() else pd.DataFrame())
decimals_for_export = int(st.session_state.get("round_decimals", 3))

# ==========================================================
# 🧾 Test Report Generation + Google Drive Upload
# ==========================================================
import io
import streamlit as st
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseUpload

def upload_to_drive(file_data, file_name):
    """Uploads the given BytesIO file to Google Drive and returns a shareable link."""
    try:
        st.write("🧠 Step 1: Starting upload_to_drive()")

        # Load credentials from Streamlit secrets
        creds_info = st.secrets["gcp_service_account"]
        folder_id = st.secrets["gdrive"]["folder_id"]
        st.write("✅ Secrets loaded successfully")

        # Create credentials
        creds = service_account.Credentials.from_service_account_info(
            creds_info,
            scopes=["https://www.googleapis.com/auth/drive.file"]
        )
        st.write("✅ Credentials object created")

        # Build the Drive service
        service = build("drive", "v3", credentials=creds)
        st.write("✅ Google Drive API client built")

        # Define metadata and media
        file_metadata = {"name": file_name, "parents": [folder_id]}
        media = MediaIoBaseUpload(file_data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.write(f"📤 Uploading file '{file_name}' to folder {folder_id}...")

        # Upload the file
        uploaded_file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id"
        ).execute()

        file_id = uploaded_file.get("id")
        st.write(f"✅ File uploaded successfully with ID: {file_id}")

        # Make the file shareable
        service.permissions().create(
            fileId=file_id,
            body={"role": "reader", "type": "anyone"}
        ).execute()
        st.write("🌍 File permissions updated for public link")

        file_link = f"https://drive.google.com/file/d/{file_id}/view?usp=sharing"
        st.success(f"✅ File uploaded: [Open in Drive]({file_link})")

        return file_link

    except Exception as e:
        st.error(f"❌ Upload failed: {e}")
        import traceback
        st.code(traceback.format_exc())
        return None




# ==========================================================
# 🎯 UI Controls
# ==========================================================
st.subheader("Test Report Actions")

out_bio = None
fname = f"{metadata_safe.get('comp_no', 'NA')}_TestReport_{metadata_safe.get('type', 'Model')}.xlsx"

# --- Generate Excel ---
if st.button("🧾 Generate Excel", key="generate_excel"):
    try:
        out_bio = generate_certificate_from_template_bytes(
            TEMPLATE_PATH,
            metadata_safe,
            perf_for_export,
            decimals=decimals_for_export
        )
        st.success("✅ Excel report generated successfully!")
    except Exception as e:
        st.error(f"Failed to generate report: {e}")

# --- Upload to Drive ---
if out_bio and st.button("☁️ Upload to Google Drive", key="upload_drive"):
    drive_link = upload_to_drive(out_bio, fname)
    if drive_link:
        st.success(f"✅ Uploaded successfully! [Open in Google Drive]({drive_link})")

# --- Download ---
if out_bio:
    st.download_button(
        "⬇️ Download Report (.xlsx)",
        data=out_bio,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_excel"
    )


